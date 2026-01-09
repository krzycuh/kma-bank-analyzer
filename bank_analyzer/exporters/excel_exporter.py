"""Excel export with formatting."""

from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Dict, Any

from bank_analyzer.utils.logger import get_logger

logger = get_logger(__name__)


class ExcelExporter:
    """Export aggregated data to Excel with formatting."""

    MONTHS_PL = {
        1: 'Sty', 2: 'Lut', 3: 'Mar', 4: 'Kwi',
        5: 'Maj', 6: 'Cze', 7: 'Lip', 8: 'Sie',
        9: 'Wrz', 10: 'Paź', 11: 'Lis', 12: 'Gru',
    }

    def export(self, aggregated_data: Dict[str, Any], output_path: Path):
        """
        Export data to Excel file.

        Args:
            aggregated_data: Aggregated data from Aggregator
            output_path: Path to output file
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.error("openpyxl not installed. Run: pip install openpyxl")
            raise ImportError("openpyxl is required for Excel export")

        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        # Backup if file exists
        if output_path.exists():
            backup_path = output_path.with_suffix(
                f'.backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            )
            output_path.rename(backup_path)
            logger.info(f"Created backup: {backup_path.name}")

        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Create sheets
        years = aggregated_data.get('years', {})
        for year in sorted(years.keys()):
            self._create_year_summary_sheet(wb, aggregated_data, year)

        uncategorized = aggregated_data.get('uncategorized', [])
        if uncategorized:
            self._create_uncategorized_sheet(wb, uncategorized)

        all_transactions = aggregated_data.get('all_transactions', [])
        if all_transactions:
            self._create_all_transactions_sheet(wb, all_transactions)

        # Save
        wb.save(output_path)
        logger.info(f"Exported to: {output_path}")

    def _create_year_summary_sheet(
        self,
        wb,
        data: Dict[str, Any],
        year: int
    ):
        """Create yearly summary sheet."""
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter

        ws = wb.create_sheet(f"Rok {year}")
        year_data = data['years'][year]

        # Title
        ws['A1'] = f"Wydatki - Rok {year}"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:N1')

        # Headers
        row = 3
        ws.cell(row, 1, "Kategoria")
        for month in range(1, 13):
            ws.cell(row, month + 1, self.MONTHS_PL[month])
        ws.cell(row, 14, "SUMA ROCZNA")

        # Format headers
        header_fill = PatternFill(
            start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
        )
        for col in range(1, 15):
            cell = ws.cell(row, col)
            cell.font = Font(bold=True)
            cell.fill = header_fill

        row += 1

        # Categories data
        categories = year_data.get('categories_year', {})

        for cat_main in sorted(categories.keys()):
            # Main category row
            start_row = row
            ws.cell(row, 1, cat_main)
            ws.cell(row, 1).font = Font(bold=True)

            # Calculate monthly sums for main category
            cat_main_monthly = [Decimal('0')] * 12

            # Subcategories
            subcats = categories[cat_main]
            for cat_sub in sorted(subcats.keys()):
                row += 1
                ws.cell(row, 1, f"  {cat_sub}")  # Indented

                # Monthly amounts
                for month in range(1, 13):
                    month_data = year_data['months'].get(month, {})
                    cats = month_data.get('categories', {})
                    amount = Decimal('0')

                    if cat_main in cats and cat_sub in cats[cat_main]:
                        amount = cats[cat_main][cat_sub].get('total', Decimal('0'))

                    if amount:
                        ws.cell(row, month + 1, float(amount))
                        ws.cell(row, month + 1).number_format = '#,##0.00'
                        cat_main_monthly[month - 1] += amount

                # Yearly total for subcategory
                total = subcats[cat_sub].get('total', Decimal('0'))
                ws.cell(row, 14, float(total))
                ws.cell(row, 14).number_format = '#,##0.00'

            # Insert main category monthly sums
            for month in range(1, 13):
                if cat_main_monthly[month - 1]:
                    ws.cell(start_row, month + 1, float(cat_main_monthly[month - 1]))
                    ws.cell(start_row, month + 1).number_format = '#,##0.00'
                    ws.cell(start_row, month + 1).font = Font(bold=True)

            # Yearly total for main category
            total_main = sum(
                s.get('total', Decimal('0')) for s in subcats.values()
            )
            ws.cell(start_row, 14, float(total_main))
            ws.cell(start_row, 14).number_format = '#,##0.00'
            ws.cell(start_row, 14).font = Font(bold=True)

            row += 1

        # Monthly sum row
        row += 1
        ws.cell(row, 1, "SUMA MIESIĘCZNA")
        ws.cell(row, 1).font = Font(bold=True)

        sum_fill = PatternFill(
            start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"
        )

        for month in range(1, 13):
            month_data = year_data['months'].get(month, {})
            total = float(month_data.get('total_expense', 0))
            if total:
                ws.cell(row, month + 1, total)
                ws.cell(row, month + 1).number_format = '#,##0.00'
                ws.cell(row, month + 1).font = Font(bold=True)
                ws.cell(row, month + 1).fill = sum_fill

        # Yearly total
        total_year = float(year_data.get('total_year_expense', 0))
        ws.cell(row, 14, total_year)
        ws.cell(row, 14).number_format = '#,##0.00'
        ws.cell(row, 14).font = Font(bold=True)
        ws.cell(row, 14).fill = sum_fill

        # Column widths
        ws.column_dimensions['A'].width = 35
        for col in range(2, 15):
            ws.column_dimensions[get_column_letter(col)].width = 12

        # Freeze panes
        ws.freeze_panes = 'B4'

    def _create_uncategorized_sheet(self, wb, uncategorized):
        """Sheet with uncategorized transactions."""
        from openpyxl.styles import Font, PatternFill

        ws = wb.create_sheet("Nieprzypisane")

        # Headers
        headers = ['Data', 'Kontrahent', 'Opis', 'Kwota', 'Bank', 'ID']
        header_fill = PatternFill(
            start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True)
            cell.fill = header_fill

        # Data
        for row, trans in enumerate(uncategorized, 2):
            ws.cell(row, 1, trans.date.strftime('%Y-%m-%d'))
            ws.cell(row, 2, trans.counterparty)
            ws.cell(row, 3, trans.description[:100])  # Limit description length
            ws.cell(row, 4, float(trans.amount))
            ws.cell(row, 5, trans.source_bank)
            ws.cell(row, 6, trans.id)

            ws.cell(row, 4).number_format = '#,##0.00'

        # Column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 18

        # AutoFilter
        if uncategorized:
            ws.auto_filter.ref = f"A1:F{len(uncategorized) + 1}"

    def _create_all_transactions_sheet(self, wb, transactions):
        """Sheet with all transactions."""
        from openpyxl.styles import Font, PatternFill

        ws = wb.create_sheet("Wszystkie transakcje")

        # Headers
        headers = [
            'Data', 'Kontrahent', 'Opis', 'Kwota',
            'Kategoria', 'Podkategoria', 'Bank', 'ID'
        ]
        header_fill = PatternFill(
            start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True)
            cell.fill = header_fill

        # Data (sorted by date descending)
        sorted_trans = sorted(transactions, key=lambda t: t.date, reverse=True)
        alt_fill = PatternFill(
            start_color="F0F0F0", end_color="F0F0F0", fill_type="solid"
        )

        for row, trans in enumerate(sorted_trans, 2):
            ws.cell(row, 1, trans.date.strftime('%Y-%m-%d'))
            ws.cell(row, 2, trans.counterparty)
            ws.cell(row, 3, trans.description[:100])
            ws.cell(row, 4, float(trans.amount))
            ws.cell(row, 5, trans.category_main or '')
            ws.cell(row, 6, trans.category_sub or '')
            ws.cell(row, 7, trans.source_bank)
            ws.cell(row, 8, trans.id)

            ws.cell(row, 4).number_format = '#,##0.00'

            # Alternate row coloring
            if row % 2 == 0:
                for col in range(1, 9):
                    ws.cell(row, col).fill = alt_fill

        # Column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 25
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 18

        # AutoFilter
        if transactions:
            ws.auto_filter.ref = f"A1:H{len(transactions) + 1}"
